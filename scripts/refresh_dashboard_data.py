from __future__ import annotations

import argparse
import base64
import datetime as dt
import html
import json
import os
import re
import subprocess
import tempfile
import time
import urllib.parse
import urllib.request
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


SCRIPT_DIR = Path(__file__).resolve().parent
BUNDLE_DIR = SCRIPT_DIR.parent
WORKBOOK_NAME_HINT = "Liseo Assemblies Master Sheet"
DEFAULT_OUTPUT = "dashboard_data.json"
SOURCE_SHEET = "Production Data"
REFERENCE_SHEET = "Reference"
CREATE_NO_WINDOW = getattr(subprocess, "CREATE_NO_WINDOW", 0)
EXCEL_MATCH_EXIT = 17

ACTIVE_TECHS = {"Innocent Mhora", "Talent Mutanda"}
MONTH_LABELS = {
    1: "Jan",
    2: "Feb",
    3: "Mar",
    4: "Apr",
    5: "May",
    6: "Jun",
    7: "Jul",
    8: "Aug",
    9: "Sep",
    10: "Oct",
    11: "Nov",
    12: "Dec",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build dashboard_data.json from the Liseo workbook.")
    parser.add_argument("--workbook", help="Local workbook path (.xlsx/.xlsm).")
    parser.add_argument("--workbook-url", help="Public share or direct download URL for the workbook.")
    parser.add_argument("--output", default=DEFAULT_OUTPUT, help="Where to write the dashboard JSON.")
    return parser.parse_args()


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def canonical_sku(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return clean_text(value)


def to_float(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return float(value)
    if isinstance(value, (int, float)):
        return float(value)
    text = clean_text(value).replace(",", "")
    if not text:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def to_int(value: Any) -> int | None:
    num = to_float(value)
    if num is None:
        return None
    return int(round(num))


def parse_date_value(value: Any) -> dt.date | None:
    if value is None or value == "":
        return None
    if isinstance(value, dt.datetime):
        return value.date()
    if isinstance(value, dt.date):
        return value
    text = clean_text(value)
    if not text:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%d-%m-%Y"):
        try:
            return dt.datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def normalize_month_label(value: Any, month_num: int | None) -> str:
    text = clean_text(value)
    if text:
        short = text[:3].title()
        if short in MONTH_LABELS.values():
            return short
    if month_num in MONTH_LABELS:
        return MONTH_LABELS[month_num]
    return ""


def find_default_workbook(bundle_dir: Path) -> Path | None:
    search_roots = [bundle_dir.parent]
    if bundle_dir.parent.parent != bundle_dir.parent:
        search_roots.append(bundle_dir.parent.parent)

    for root in search_roots:
        if not root.exists():
            continue

        preferred: list[Path] = []
        for pattern in ("*.xlsm", "*.xlsx"):
            preferred.extend(
                sorted(path for path in root.glob(pattern) if WORKBOOK_NAME_HINT.lower() in path.name.lower())
            )
        if preferred:
            return preferred[0]

        for pattern in ("*.xlsm", "*.xlsx"):
            matches = sorted(root.glob(pattern))
            if matches:
                return matches[0]
    return None


def with_download_hint(url: str) -> str:
    parsed = urllib.parse.urlsplit(url)
    query = urllib.parse.parse_qsl(parsed.query, keep_blank_values=True)
    keys = {key.lower() for key, _ in query}
    if "download" not in keys:
        query.append(("download", "1"))
    return urllib.parse.urlunsplit((parsed.scheme, parsed.netloc, parsed.path, urllib.parse.urlencode(query), parsed.fragment))


def candidate_urls(url: str) -> list[str]:
    candidates = [url]
    hinted = with_download_hint(url)
    if hinted != url:
        candidates.append(hinted)
    return candidates


def extract_download_url_from_html(page_html: str) -> str | None:
    patterns = [
        r'"downloadUrl":"([^"]+)"',
        r'"@content\.downloadUrl":"([^"]+)"',
        r'"downloadUrl\\":\\"([^"]+)\\"',
    ]
    for pattern in patterns:
        match = re.search(pattern, page_html)
        if not match:
            continue
        candidate = match.group(1)
        candidate = candidate.replace("\\u0026", "&").replace("\\/", "/").replace('\\"', '"')
        return html.unescape(candidate)
    return None


def share_token(url: str) -> str:
    raw = base64.b64encode(url.encode("utf-8")).decode("ascii").rstrip("=")
    return "u!" + raw.replace("/", "_").replace("+", "-")


def request_json(url: str, headers: dict[str, str] | None = None, data: bytes | None = None) -> dict[str, Any]:
    request = urllib.request.Request(url, headers=headers or {}, data=data)
    with urllib.request.urlopen(request, timeout=60) as response:
        return json.loads(response.read().decode("utf-8"))


def request_bytes(url: str, headers: dict[str, str] | None = None) -> bytes:
    request = urllib.request.Request(url, headers=headers or {})
    with urllib.request.urlopen(request, timeout=60) as response:
        return response.read()


def find_download_url(payload: Any) -> str | None:
    if isinstance(payload, dict):
        for key, value in payload.items():
            if "downloadurl" in key.lower() and isinstance(value, str) and value.startswith("http"):
                return value
            found = find_download_url(value)
            if found:
                return found
    elif isinstance(payload, list):
        for item in payload:
            found = find_download_url(item)
            if found:
                return found
    return None


def write_temp_workbook(payload: bytes, suffix: str = ".xlsm") -> Path:
    fd, temp_path = tempfile.mkstemp(suffix=suffix)
    os.close(fd)
    path = Path(temp_path)
    path.write_bytes(payload)
    return path


def ensure_excel_file(path: Path) -> None:
    if zipfile.is_zipfile(path):
        return
    raise ValueError(f"Downloaded file is not a valid Excel workbook: {path.name}")


def guessed_name(url: str, headers: Any) -> str:
    content_disposition = headers.get("Content-Disposition", "")
    if "filename=" in content_disposition:
        filename = content_disposition.split("filename=", 1)[1].strip().strip('"')
        if filename:
            return filename

    parsed = urllib.parse.urlsplit(url)
    filename = Path(parsed.path).name
    if filename:
        return filename
    return "downloaded_workbook.xlsm"


def onedrive_badger_headers() -> dict[str, str]:
    token_payload = request_json(
        "https://api-badgerp.svc.ms/v1.0/token",
        headers={"Content-Type": "application/json", "User-Agent": "Mozilla/5.0"},
        data=json.dumps({"appId": "5cbed6ac-a083-4e14-b191-b4ba07653de2"}).encode("utf-8"),
    )
    token = token_payload.get("token")
    if not token:
        raise RuntimeError("Could not get OneDrive public access token.")
    return {
        "Authorization": f"Badger {token}",
        "Prefer": "autoredeem",
        "User-Agent": "Mozilla/5.0",
        "Accept": "application/json, */*",
    }


def download_onedrive_share(url: str) -> tuple[Path, str]:
    token = share_token(url)
    headers = onedrive_badger_headers()
    metadata = request_json(f"https://api.onedrive.com/v1.0/shares/{token}/driveItem", headers=headers)

    download_url = find_download_url(metadata)
    if download_url:
        data = request_bytes(download_url, headers={"User-Agent": "Mozilla/5.0"})
        filename = metadata.get("name") or guessed_name(download_url, {})
        suffix = Path(filename).suffix or ".xlsm"
        target = write_temp_workbook(data, suffix=suffix)
        ensure_excel_file(target)
        return target, filename

    for content_url in (
        f"https://api.onedrive.com/v1.0/shares/{token}/driveItem/content",
        f"https://api.onedrive.com/v1.0/shares/{token}/root/content",
    ):
        try:
            data = request_bytes(content_url, headers=headers)
            filename = metadata.get("name") or "downloaded_workbook.xlsm"
            suffix = Path(filename).suffix or ".xlsm"
            target = write_temp_workbook(data, suffix=suffix)
            ensure_excel_file(target)
            return target, filename
        except Exception:  # noqa: BLE001
            continue

    raise RuntimeError("OneDrive share metadata loaded, but no downloadable workbook URL was available.")


def download_workbook(url: str) -> tuple[Path, str]:
    headers = {"User-Agent": "Mozilla/5.0 Codex Dashboard Refresher"}

    last_error: Exception | None = None
    for candidate in candidate_urls(url):
        try:
            request = urllib.request.Request(candidate, headers=headers)
            with urllib.request.urlopen(request, timeout=60) as response:
                payload = response.read()
                final_url = response.geturl()
                filename = guessed_name(final_url, response.headers)
                suffix = Path(filename).suffix or ".xlsm"
                content_type = response.headers.get("Content-Type", "")
                if payload.startswith(b"PK\x03\x04"):
                    target = write_temp_workbook(payload, suffix=suffix)
                    ensure_excel_file(target)
                    return target, filename
                if "html" in content_type.lower():
                    nested_url = extract_download_url_from_html(payload.decode("utf-8", errors="ignore"))
                    if nested_url:
                        nested_request = urllib.request.Request(nested_url, headers={"User-Agent": "Mozilla/5.0"})
                        with urllib.request.urlopen(nested_request, timeout=60) as nested_response:
                            nested_payload = nested_response.read()
                            if nested_payload.startswith(b"PK\x03\x04"):
                                nested_final_url = nested_response.geturl()
                                nested_name = guessed_name(nested_final_url, nested_response.headers)
                                nested_suffix = Path(nested_name).suffix or ".xlsm"
                                target = write_temp_workbook(nested_payload, suffix=nested_suffix)
                                ensure_excel_file(target)
                                return target, nested_name
                raise RuntimeError(f"Workbook URL did not return an Excel file: {candidate}")
        except Exception as exc:  # noqa: BLE001
            last_error = exc

    url_lower = url.lower()
    if "1drv.ms" in url_lower or "onedrive.live.com" in url_lower:
        try:
            return download_onedrive_share(url)
        except Exception as exc:  # noqa: BLE001
            last_error = exc

    raise RuntimeError(f"Unable to download workbook from {url}") from last_error


def create_excel_snapshot(workbook_path: Path) -> Path:
    helper_script = Path(__file__).with_name("save_excel_snapshot.ps1")
    if not helper_script.exists():
        raise FileNotFoundError(f"Snapshot helper not found: {helper_script}")

    fd, temp_path = tempfile.mkstemp(suffix=workbook_path.suffix or ".xlsx")
    os.close(fd)
    snapshot_path = Path(temp_path)

    startupinfo = None
    if os.name == "nt":
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = 0

    command = [
        "powershell",
        "-NoProfile",
        "-ExecutionPolicy",
        "Bypass",
        "-File",
        str(helper_script),
        "-SourcePath",
        str(workbook_path),
        "-TargetPath",
        str(snapshot_path),
    ]
    last_message = "Excel could not create a readable snapshot."
    for attempt in range(4):
        result = subprocess.run(
            command,
            check=False,
            text=True,
            capture_output=True,
            creationflags=CREATE_NO_WINDOW,
            startupinfo=startupinfo,
        )
        if result.returncode == 0:
            return snapshot_path

        last_message = result.stderr.strip() or result.stdout.strip() or last_message
        snapshot_path.unlink(missing_ok=True)
        if "0x800AC472" not in last_message or attempt == 3:
            break
        time.sleep(1.5)

    raise RuntimeError(last_message)


def workbook_lockfile_path(workbook_path: Path) -> Path:
    return workbook_path.with_name(f"~${workbook_path.name}")


def is_workbook_open_in_excel(workbook_path: Path) -> bool:
    if os.name != "nt":
        return False

    if workbook_lockfile_path(workbook_path).exists():
        return True

    escaped_path = str(workbook_path.resolve()).replace("'", "''")
    command = rf"""
$path = '{escaped_path}'
try {{
    $excel = [Runtime.InteropServices.Marshal]::GetActiveObject('Excel.Application')
}} catch {{
    exit 0
}}

$workbooks = $excel.Workbooks
for ($index = 1; $index -le $workbooks.Count; $index++) {{
    $candidate = $workbooks.Item($index)
    if ($candidate.FullName -and [string]::Equals([System.IO.Path]::GetFullPath($candidate.FullName), $path, [System.StringComparison]::OrdinalIgnoreCase)) {{
        exit {EXCEL_MATCH_EXIT}
    }}
}}

exit 0
"""

    startupinfo = None
    if os.name == "nt":
        startupinfo = subprocess.STARTUPINFO()
        startupinfo.dwFlags |= subprocess.STARTF_USESHOWWINDOW
        startupinfo.wShowWindow = 0

    result = subprocess.run(
        [
            "powershell",
            "-NoProfile",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            command,
        ],
        check=False,
        text=True,
        capture_output=True,
        creationflags=CREATE_NO_WINDOW,
        startupinfo=startupinfo,
    )
    return result.returncode == EXCEL_MATCH_EXIT


def load_dashboard_workbook(workbook_path: Path) -> tuple[Any, Path | None]:
    if is_workbook_open_in_excel(workbook_path):
        try:
            snapshot_path = create_excel_snapshot(workbook_path)
            return load_workbook(snapshot_path, data_only=True, read_only=True), snapshot_path
        except Exception:
            pass

    try:
        return load_workbook(workbook_path, data_only=True, read_only=True), None
    except PermissionError:
        snapshot_path = create_excel_snapshot(workbook_path)
        return load_workbook(snapshot_path, data_only=True, read_only=True), snapshot_path


def cleanup_temp_file(path: Path | None) -> None:
    if path is None:
        return
    for _ in range(6):
        try:
            path.unlink(missing_ok=True)
            return
        except PermissionError:
            time.sleep(0.5)


def resolve_workbook_source(workbook: str | None, workbook_url: str | None) -> tuple[Path, bool, str]:
    if workbook:
        path = Path(workbook).expanduser().resolve()
        if not path.exists():
            raise FileNotFoundError(f"Workbook not found: {path}")
        return path, False, path.name

    env_path = os.getenv("WORKBOOK_PATH")
    if env_path:
        path = Path(env_path).expanduser().resolve()
        if path.exists():
            return path, False, path.name

    url = workbook_url or os.getenv("WORKBOOK_URL")
    if url:
        path, workbook_name = download_workbook(url)
        return path, True, workbook_name

    default_path = find_default_workbook(BUNDLE_DIR)
    if default_path is not None:
        resolved = default_path.resolve()
        return resolved, False, resolved.name

    raise ValueError("Provide --workbook, WORKBOOK_PATH, --workbook-url, or WORKBOOK_URL.")


def load_reference_map(workbook: Any) -> dict[str, dict[str, Any]]:
    ws = workbook[REFERENCE_SHEET]
    ref_map: dict[str, dict[str, Any]] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku = canonical_sku(row[0])
        if not sku:
            continue
        ref_map[sku] = {
            "model": clean_text(row[1]),
            "difficulty": to_float(row[2]) or 0.0,
            "rate": to_float(row[3]) or 0.0,
        }
    return ref_map


def build_rows(workbook: Any, ref_map: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    ws = workbook[SOURCE_SHEET]
    rows: list[dict[str, Any]] = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        qc_date = parse_date_value(row[1])
        qty = to_int(row[5])
        sku = canonical_sku(row[6])
        tech = clean_text(row[8])

        if not qc_date and qty is None and not sku and not tech:
            continue
        if qty is None or qty <= 0 or not tech:
            continue

        year = to_int(row[2]) or (qc_date.year if qc_date else None)
        month_num = to_int(row[4]) or (qc_date.month if qc_date else None)
        if year is None or month_num is None:
            continue

        month_label = normalize_month_label(row[3], month_num)
        ref = ref_map.get(sku, {})
        difficulty = to_float(row[10])
        if difficulty in (None, 0.0):
            difficulty = float(ref.get("difficulty", 0.0))

        total_points = to_float(row[11])
        if total_points is None:
            total_points = round(qty * difficulty, 2)

        status = clean_text(row[9]) or ("Active" if tech in ACTIVE_TECHS else "Inactive")
        rows.append(
            {
                "tech": tech,
                "yr": year,
                "mo": month_label,
                "moNum": month_num,
                "u": qty,
                "p": round(total_points, 1),
                "s": status,
            }
        )

    return rows


def build_payload(rows: list[dict[str, Any]], workbook_name: str, source_modified_at: dt.datetime | None) -> dict[str, Any]:
    generated_at = dt.datetime.now(dt.timezone.utc)
    total_units = sum(int(row["u"]) for row in rows)
    total_points = round(sum(float(row["p"]) for row in rows), 1)
    techs = sorted({row["tech"] for row in rows})
    active_techs = sorted({row["tech"] for row in rows if row["s"] == "Active"})

    return {
        "meta": {
            "workbookName": workbook_name,
            "generatedAt": generated_at.isoformat(),
            "generatedAtLabel": generated_at.strftime("%d %b %Y %H:%M UTC"),
            "sourceModifiedAt": source_modified_at.isoformat() if source_modified_at else None,
            "sourceSheet": SOURCE_SHEET,
        },
        "stats": {
            "rows": len(rows),
            "totalUnits": total_units,
            "totalPoints": total_points,
            "techCount": len(techs),
            "activeTechCount": len(active_techs),
        },
        "rows": rows,
    }


def refresh_dashboard_data(workbook: str | None = None, workbook_url: str | None = None, output: str | Path = DEFAULT_OUTPUT) -> Path:
    workbook_path, is_temp, workbook_name = resolve_workbook_source(workbook, workbook_url)
    workbook_obj = None
    snapshot_path: Path | None = None
    try:
        workbook_obj, snapshot_path = load_dashboard_workbook(workbook_path)
        ref_map = load_reference_map(workbook_obj)
        rows = build_rows(workbook_obj, ref_map)
        source_modified_at = dt.datetime.fromtimestamp(workbook_path.stat().st_mtime, dt.timezone.utc)
        payload = build_payload(rows, workbook_name, source_modified_at)
        output_path = Path(output).expanduser().resolve()
        output_path.write_text(json.dumps(payload, indent=2), encoding="utf-8")
        return output_path
    finally:
        if workbook_obj is not None:
            workbook_obj.close()
        if snapshot_path is not None:
            cleanup_temp_file(snapshot_path)
        if is_temp:
            cleanup_temp_file(workbook_path)


def main() -> None:
    args = parse_args()
    output_path = refresh_dashboard_data(
        workbook=args.workbook,
        workbook_url=args.workbook_url,
        output=args.output,
    )
    print(output_path)


if __name__ == "__main__":
    main()
